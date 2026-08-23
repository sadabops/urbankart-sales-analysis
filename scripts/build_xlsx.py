import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

df = pd.read_csv('../data/urbankart_sales_clean.csv')
n = len(df)  # 4820

wb = Workbook()

# ---------- Styles ----------
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E78')
SUBTITLE_FONT = Font(name='Arial', italic=True, size=10, color='666666')
LABEL_FONT = Font(name='Arial', bold=True, size=11)
NORMAL_FONT = Font(name='Arial', size=11)
ANSWER_FILL = PatternFill('solid', fgColor='E2EFDA')
ANSWER_FONT = Font(name='Arial', bold=True, size=12, color='375623')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ================= Sheet 1: Data =================
ws = wb.active
ws.title = 'Data'
cols = list(df.columns)
for j, col in enumerate(cols, start=1):
    ws.cell(row=1, column=j, value=col)
style_header_row(ws, 1, len(cols))

for i, row in enumerate(df.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

ws.freeze_panes = 'A2'
autosize(ws, [16, 12, 10, 12, 26, 16, 9, 10, 13, 12, 13])
last_row = n + 1

# ================= Sheet 2: Q1 Region Revenue =================
ws2 = wb.create_sheet('Q1_Region_Revenue')
ws2['A1'] = 'Q1 — Which region is carrying us?'
ws2['A1'].font = TITLE_FONT
ws2['A2'] = 'Total revenue by region, Jan–Jun 2026 (from Data sheet)'
ws2['A2'].font = SUBTITLE_FONT

headers = ['Region', 'Total Revenue (₹)', 'Share of Total']
for j, h in enumerate(headers, start=1):
    ws2.cell(row=4, column=j, value=h)
style_header_row(ws2, 4, len(headers))

regions = ['South', 'North', 'West', 'East', 'Central']
start_r = 5
for i, region in enumerate(regions):
    r = start_r + i
    ws2.cell(row=r, column=1, value=region).font = NORMAL_FONT
    ws2.cell(row=r, column=2, value=f"=SUMIFS(Data!$I$2:$I${last_row},Data!$C$2:$C${last_row},A{r})")
    ws2.cell(row=r, column=2).number_format = '#,##0'
    ws2.cell(row=r, column=3, value=f"=B{r}/SUM($B${start_r}:$B${start_r+len(regions)-1})")
    ws2.cell(row=r, column=3).number_format = '0.0%'
    for c in range(1,4):
        ws2.cell(row=r, column=c).border = BORDER

total_r = start_r + len(regions)
ws2.cell(row=total_r, column=1, value='Total').font = LABEL_FONT
ws2.cell(row=total_r, column=2, value=f"=SUM(B{start_r}:B{start_r+len(regions)-1})")
ws2.cell(row=total_r, column=2).number_format = '#,##0'
ws2.cell(row=total_r, column=2).font = LABEL_FONT
ws2.cell(row=total_r, column=3, value=f"=SUM(C{start_r}:C{start_r+len(regions)-1})")
ws2.cell(row=total_r, column=3).number_format = '0.0%'
ws2.cell(row=total_r, column=3).font = LABEL_FONT

ws2['A11'] = 'Answer:'
ws2['A11'].font = LABEL_FONT
ws2['B11'] = f"=INDEX(A{start_r}:A{start_r+len(regions)-1},MATCH(MAX(B{start_r}:B{start_r+len(regions)-1}),B{start_r}:B{start_r+len(regions)-1},0)) & \" — ₹\" & TEXT(MAX(B{start_r}:B{start_r+len(regions)-1}),\"#,##0\")"
ws2['B11'].font = ANSWER_FONT
ws2['B11'].fill = ANSWER_FILL

autosize(ws2, [14, 20, 16])

# chart
chart = BarChart()
chart.title = "Total Revenue by Region"
chart.y_axis.title = 'Revenue (₹)'
chart.x_axis.title = 'Region'
data_ref = Reference(ws2, min_col=2, min_row=4, max_row=start_r+len(regions)-1)
cats_ref = Reference(ws2, min_col=1, min_row=start_r, max_row=start_r+len(regions)-1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 16
chart.height = 9
ws2.add_chart(chart, 'E4')

# ================= Sheet 3: Q2 Monthly Revenue =================
ws3 = wb.create_sheet('Q2_Monthly_Revenue')
ws3['A1'] = 'Q2 — Revenue momentum (Month-on-Month)'
ws3['A1'].font = TITLE_FONT
ws3['A2'] = 'MoM % = (This month − Last month) ÷ Last month × 100'
ws3['A2'].font = SUBTITLE_FONT

headers = ['Month', 'Total Revenue (₹)', 'MoM % Change']
for j, h in enumerate(headers, start=1):
    ws3.cell(row=4, column=j, value=h)
style_header_row(ws3, 4, len(headers))

months = ['2026-01','2026-02','2026-03','2026-04','2026-05','2026-06']
month_labels = ['Jan 2026','Feb 2026','Mar 2026','Apr 2026','May 2026','Jun 2026']
start_r3 = 5
for i, (m, lbl) in enumerate(zip(months, month_labels)):
    r = start_r3 + i
    ws3.cell(row=r, column=1, value=lbl).font = NORMAL_FONT
    # LEFT(order_date,7) matches YYYY-MM prefix
    ws3.cell(row=r, column=2, value=f'=SUMPRODUCT((LEFT(Data!$B$2:$B${last_row},7)="{m}")*Data!$I$2:$I${last_row})')
    ws3.cell(row=r, column=2).number_format = '#,##0'
    if i == 0:
        ws3.cell(row=r, column=3, value='—')
    else:
        ws3.cell(row=r, column=3, value=f'=(B{r}-B{r-1})/B{r-1}')
        ws3.cell(row=r, column=3).number_format = '0.0%'
    for c in range(1,4):
        ws3.cell(row=r, column=c).border = BORDER

ws3['A12'] = 'Apr → May:'
ws3['A12'].font = LABEL_FONT
ws3['B12'] = '=C9'
ws3['B12'].number_format = '0.0%'
ws3['B12'].font = ANSWER_FONT
ws3['B12'].fill = ANSWER_FILL

ws3['A13'] = 'May → Jun:'
ws3['A13'].font = LABEL_FONT
ws3['B13'] = '=C10'
ws3['B13'].number_format = '0.0%'
ws3['B13'].font = ANSWER_FONT
ws3['B13'].fill = ANSWER_FILL

ws3['A14'] = 'Direction:'
ws3['A14'].font = LABEL_FONT
ws3['B14'] = '=IF(AND(B12>0,B13>0),"Growing",IF(AND(B12<0,B13<0),"Declining","Mixed"))'
ws3['B14'].font = ANSWER_FONT
ws3['B14'].fill = ANSWER_FILL

autosize(ws3, [12, 20, 16])

chart2 = LineChart()
chart2.title = "Monthly Revenue Trend"
chart2.y_axis.title = 'Revenue (₹)'
chart2.x_axis.title = 'Month'
data_ref2 = Reference(ws3, min_col=2, min_row=4, max_row=start_r3+5)
cats_ref2 = Reference(ws3, min_col=1, min_row=start_r3, max_row=start_r3+5)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
chart2.width = 16
chart2.height = 9
ws3.add_chart(chart2, 'E4')

# ================= Sheet 4: Q3 Channel Analysis =================
ws4 = wb.create_sheet('Q3_Channel_Analysis')
ws4['A1'] = 'Q3 — The channel decision'
ws4['A1'].font = TITLE_FONT
ws4['A2'] = 'Looking beyond total revenue: trajectory (growth) and order economics (AOV)'
ws4['A2'].font = SUBTITLE_FONT

headers = ['Channel', 'Total Revenue (₹)', 'Share of Total', 'Orders', 'Avg Order Value (₹)', 'Jan Revenue', 'Jun Revenue', 'Jan→Jun Growth %']
for j, h in enumerate(headers, start=1):
    ws4.cell(row=4, column=j, value=h)
style_header_row(ws4, 4, len(headers))

channels = ['App', 'Marketplace', 'Website']
start_r4 = 5
for i, ch in enumerate(channels):
    r = start_r4 + i
    ws4.cell(row=r, column=1, value=ch).font = NORMAL_FONT
    ws4.cell(row=r, column=2, value=f"=SUMIFS(Data!$I$2:$I${last_row},Data!$K$2:$K${last_row},A{r})")
    ws4.cell(row=r, column=2).number_format = '#,##0'
    ws4.cell(row=r, column=3, value=f"=B{r}/SUM($B${start_r4}:$B${start_r4+2})")
    ws4.cell(row=r, column=3).number_format = '0.0%'
    ws4.cell(row=r, column=4, value=f"=COUNTIFS(Data!$K$2:$K${last_row},A{r})")
    ws4.cell(row=r, column=5, value=f"=B{r}/D{r}")
    ws4.cell(row=r, column=5).number_format = '#,##0'
    ws4.cell(row=r, column=6, value=f'=SUMPRODUCT((LEFT(Data!$B$2:$B${last_row},7)="2026-01")*(Data!$K$2:$K${last_row}=A{r})*Data!$I$2:$I${last_row})')
    ws4.cell(row=r, column=6).number_format = '#,##0'
    ws4.cell(row=r, column=7, value=f'=SUMPRODUCT((LEFT(Data!$B$2:$B${last_row},7)="2026-06")*(Data!$K$2:$K${last_row}=A{r})*Data!$I$2:$I${last_row})')
    ws4.cell(row=r, column=7).number_format = '#,##0'
    ws4.cell(row=r, column=8, value=f"=(G{r}-F{r})/F{r}")
    ws4.cell(row=r, column=8).number_format = '0.0%'
    for c in range(1,9):
        ws4.cell(row=r, column=c).border = BORDER

ws4['A9'] = 'Part A — Weakest channel by revenue:'
ws4['A9'].font = LABEL_FONT
ws4['C9'] = f'=INDEX(A{start_r4}:A{start_r4+2},MATCH(MIN(B{start_r4}:B{start_r4+2}),B{start_r4}:B{start_r4+2},0)) & " — " & TEXT(INDEX(C{start_r4}:C{start_r4+2},MATCH(MIN(B{start_r4}:B{start_r4+2}),B{start_r4}:B{start_r4+2},0)),"0.0%") & " of total revenue"'
ws4['C9'].font = ANSWER_FONT
ws4['C9'].fill = ANSWER_FILL
ws4.merge_cells('C9:H9')

ws4['A11'] = 'Part B — Recommendation:'
ws4['A11'].font = LABEL_FONT
ws4['C11'] = 'DO NOT PAUSE Marketplace (see memo tab)'
ws4['C11'].font = ANSWER_FONT
ws4['C11'].fill = ANSWER_FILL
ws4.merge_cells('C11:H11')

ws4['A13'] = 'Key evidence beyond total revenue:'
ws4['A13'].font = LABEL_FONT
ws4['A14'] = '• Marketplace grew fastest Jan→Jun (see column H) — it is gaining share, not shrinking'
ws4['A15'] = '• Marketplace has the highest Avg Order Value (see column E) — most valuable order economics'
ws4['A16'] = '• Website revenue fell in the most recent month (see Q2 monthly trend by channel below)'
for rr in [14,15,16]:
    ws4.cell(row=rr, column=1).font = NORMAL_FONT

autosize(ws4, [14, 18, 14, 10, 18, 14, 14, 16])

chart3 = BarChart()
chart3.title = "Channel Revenue vs AOV"
data_ref3 = Reference(ws4, min_col=2, min_row=4, max_row=start_r4+2)
cats_ref3 = Reference(ws4, min_col=1, min_row=start_r4, max_row=start_r4+2)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
chart3.width = 16
chart3.height = 9
ws4.add_chart(chart3, 'A19')

# ================= Sheet 5: Channel Monthly Trend (supporting) =================
ws5 = wb.create_sheet('Channel_Monthly_Trend')
ws5['A1'] = 'Supporting data — Monthly revenue by channel'
ws5['A1'].font = TITLE_FONT
headers = ['Month'] + channels
for j, h in enumerate(headers, start=1):
    ws5.cell(row=3, column=j, value=h)
style_header_row(ws5, 3, len(headers))

start_r5 = 4
for i, (m, lbl) in enumerate(zip(months, month_labels)):
    r = start_r5 + i
    ws5.cell(row=r, column=1, value=lbl).font = NORMAL_FONT
    for ci, ch in enumerate(channels, start=2):
        col_letter = get_column_letter(ci)
        ws5.cell(row=r, column=ci, value=f'=SUMPRODUCT((LEFT(Data!$B$2:$B${last_row},7)="{m}")*(Data!$K$2:$K${last_row}="{ch}")*Data!$I$2:$I${last_row})')
        ws5.cell(row=r, column=ci).number_format = '#,##0'
    for c in range(1, len(headers)+1):
        ws5.cell(row=r, column=c).border = BORDER

autosize(ws5, [12, 14, 14, 14])

chart4 = LineChart()
chart4.title = "Monthly Revenue by Channel"
chart4.y_axis.title = 'Revenue (₹)'
data_ref4 = Reference(ws5, min_col=2, max_col=4, min_row=3, max_row=start_r5+5)
cats_ref4 = Reference(ws5, min_col=1, min_row=start_r5, max_row=start_r5+5)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
chart4.width = 18
chart4.height = 10
ws5.add_chart(chart4, 'F3')

# ================= Sheet 6: Summary (Answers) =================
ws6 = wb.create_sheet('Summary_Answers')
ws6.sheet_view.showGridLines = False
ws6['B2'] = 'UrbanKart — Board Meeting Prep: Answer Summary'
ws6['B2'].font = Font(name='Arial', bold=True, size=16, color='1F4E78')
ws6.merge_cells('B2:F2')
ws6['B3'] = 'Prepared for Rajesh Malhotra (COO) — Jan–Jun 2026 data'
ws6['B3'].font = SUBTITLE_FONT
ws6.merge_cells('B3:F3')

ws6['B5'] = 'Q1: Top region'
ws6['B5'].font = LABEL_FONT
ws6['C5'] = "='Q1_Region_Revenue'!B11"
ws6['C5'].font = ANSWER_FONT
ws6['C5'].fill = ANSWER_FILL
ws6.merge_cells('C5:F5')

ws6['B7'] = 'Q2: April → May'
ws6['B7'].font = LABEL_FONT
ws6['C7'] = "='Q2_Monthly_Revenue'!B12"
ws6['C7'].number_format = '0.0%'
ws6['C7'].font = ANSWER_FONT
ws6['C7'].fill = ANSWER_FILL

ws6['B8'] = 'Q2: May → June'
ws6['B8'].font = LABEL_FONT
ws6['C8'] = "='Q2_Monthly_Revenue'!B13"
ws6['C8'].number_format = '0.0%'
ws6['C8'].font = ANSWER_FONT
ws6['C8'].fill = ANSWER_FILL

ws6['B9'] = 'Q2: Direction'
ws6['B9'].font = LABEL_FONT
ws6['C9'] = "='Q2_Monthly_Revenue'!B14"
ws6['C9'].font = ANSWER_FONT
ws6['C9'].fill = ANSWER_FILL

ws6['B11'] = 'Q3a: Weakest channel by revenue'
ws6['B11'].font = LABEL_FONT
ws6['C11'] = "='Q3_Channel_Analysis'!C9"
ws6['C11'].font = ANSWER_FONT
ws6['C11'].fill = ANSWER_FILL
ws6.merge_cells('C11:F11')

ws6['B13'] = 'Q3b: Recommendation'
ws6['B13'].font = LABEL_FONT
ws6['C13'] = 'Do NOT pause Marketplace — see Memo tab for full reasoning'
ws6['C13'].font = ANSWER_FONT
ws6['C13'].fill = ANSWER_FILL
ws6.merge_cells('C13:F13')

autosize(ws6, [4, 30, 40, 14, 14, 14])

# reorder: put Summary first
wb._sheets = [ws6, ws, ws2, ws3, ws4, ws5]
wb.active = 0

wb.save('../deliverables/UrbanKart_Board_Analysis.xlsx')
print("saved")
